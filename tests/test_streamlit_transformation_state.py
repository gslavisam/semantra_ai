"""Tests Streamlit transformation-related session state and helper behavior."""

from __future__ import annotations

import ast
import json
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook


STREAMLIT_APP_PATH = Path(__file__).resolve().parents[1] / "streamlit_app.py"


def load_streamlit_functions(*function_names: str):
    source = STREAMLIT_APP_PATH.read_text(encoding="utf-8")
    module = ast.parse(source, filename=str(STREAMLIT_APP_PATH))
    selected_nodes = [
        node for node in module.body if isinstance(node, ast.FunctionDef) and node.name in function_names
    ]
    fake_streamlit = SimpleNamespace(session_state={})
    namespace = {
        "st": fake_streamlit,
        "DEFAULT_API_BASE_URL": "http://127.0.0.1:8000",
    }
    exec(compile(ast.Module(body=selected_nodes, type_ignores=[]), str(STREAMLIT_APP_PATH), "exec"), namespace)
    return fake_streamlit, [namespace[name] for name in function_names]


def test_resolve_suggested_transformation_code_drops_stale_code_after_target_change() -> None:
    _, [resolve_suggested_transformation_code] = load_streamlit_functions("resolve_suggested_transformation_code")

    resolved = resolve_suggested_transformation_code(
        {
            "target": "full_name",
            "suggested_target": "customer_name",
            "suggested_transformation_code": 'df_source["email"].str.title()',
        },
        'df_source["email"]',
    )

    assert resolved == ""


def test_build_mapping_decisions_excludes_stale_suggested_transformation_code() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "build_mapping_decisions",
    )
    build_mapping_decisions = functions[-1]

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "email": {
                    "target": "full_name",
                    "status": "accepted",
                    "suggested_target": "customer_name",
                    "suggested_transformation_code": 'df_source["email"].str.split("@").str[0].str.title()',
                }
            },
            "transform_email": True,
            "manual_transform_email": "",
            "manual_apply_email": False,
        }
    )

    assert build_mapping_decisions() == [
        {
            "source": "email",
            "target": "full_name",
            "status": "accepted",
            "resolution_type": "direct_mapping",
        }
    ]


def test_build_mapping_decisions_keeps_suggested_transformation_for_original_target() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "build_mapping_decisions",
    )
    build_mapping_decisions = functions[-1]

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "email": {
                    "target": "customer_name",
                    "status": "accepted",
                    "suggested_target": "customer_name",
                    "suggested_transformation_code": 'df_source["email"].str.split("@").str[0].str.title()',
                }
            },
            "transform_email": True,
            "manual_transform_email": "",
            "manual_apply_email": False,
        }
    )

    assert build_mapping_decisions() == [
        {
            "source": "email",
            "target": "customer_name",
            "status": "accepted",
            "resolution_type": "direct_mapping",
            "transformation_code": 'df_source["email"].str.split("@").str[0].str.title()',
        }
    ]


def test_build_mapping_decisions_preserves_non_default_resolution_type() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "build_mapping_decisions",
    )
    build_mapping_decisions = functions[-1]

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "op_type": {
                    "target": "operation_category",
                    "status": "accepted",
                    "resolution_type": "fixed_value",
                }
            },
            "manual_transform_op_type": "",
            "manual_apply_op_type": False,
        }
    )

    assert build_mapping_decisions() == [
        {
            "source": "op_type",
            "target": "operation_category",
            "status": "accepted",
            "resolution_type": "fixed_value",
        }
    ]


def test_build_mapping_decisions_includes_resolution_payload_for_fixed_value() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "build_mapping_decisions",
    )
    build_mapping_decisions = functions[-1]

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "op_type": {
                    "target": "operation_category",
                    "status": "accepted",
                    "resolution_type": "fixed_value",
                    "resolution_payload": {"value": "CUSTOMER"},
                }
            },
            "manual_transform_op_type": "",
            "manual_apply_op_type": False,
        }
    )

    assert build_mapping_decisions() == [
        {
            "source": "op_type",
            "target": "operation_category",
            "status": "accepted",
            "resolution_type": "fixed_value",
            "resolution_payload": {"value": "CUSTOMER"},
        }
    ]


def test_build_mapping_decisions_includes_reason_for_target_managed() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "build_mapping_decisions",
    )
    build_mapping_decisions = functions[-1]

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "managed_id": {
                    "target": "employee_id",
                    "status": "accepted",
                    "resolution_type": "target_managed",
                    "resolution_payload": {"reason": "Generated by destination system."},
                }
            },
            "manual_transform_managed_id": "",
            "manual_apply_managed_id": False,
        }
    )

    assert build_mapping_decisions() == [
        {
            "source": "managed_id",
            "target": "employee_id",
            "status": "accepted",
            "resolution_type": "target_managed",
            "resolution_payload": {"reason": "Generated by destination system."},
        }
    ]


def test_effective_transformation_helpers_tolerate_none_manual_code() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "effective_transformation_code",
        "transformation_mode",
    )
    effective_transformation_code, transformation_mode = functions

    fake_streamlit.session_state.update(
        {
            "manual_transform_email": None,
            "manual_apply_email": True,
            "transform_email": True,
        }
    )

    assert effective_transformation_code("email", 'df_source["email"]') == 'df_source["email"]'
    assert transformation_mode("email", 'df_source["email"]') == "suggested"


def test_build_pending_corrections_includes_rejected_review_decision() -> None:
    fake_streamlit, [build_pending_corrections] = load_streamlit_functions("build_pending_corrections")

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "cust_ref": {
                    "target": "customer_id",
                    "suggested_target": "customer_id",
                    "status": "rejected",
                }
            }
        }
    )

    assert build_pending_corrections() == [
        {
            "source": "cust_ref",
            "suggested_target": "customer_id",
            "corrected_target": None,
            "status": "rejected",
        }
    ]


def test_build_pending_corrections_skips_needs_review_override() -> None:
    fake_streamlit, [build_pending_corrections] = load_streamlit_functions("build_pending_corrections")

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "phone": {
                    "target": "phone_number",
                    "suggested_target": "contact_phone",
                    "status": "needs_review",
                }
            }
        }
    )

    assert build_pending_corrections() == []


def test_correction_governance_block_reason_reports_unclosed_review_status() -> None:
    fake_streamlit, [correction_governance_block_reason] = load_streamlit_functions("correction_governance_block_reason")

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "phone": {
                    "target": "phone_number",
                    "suggested_target": "contact_phone",
                    "status": "needs_review",
                }
            }
        }
    )

    assert correction_governance_block_reason() == (
        "Saving reviewed corrections is blocked until pending corrections come from closed review outcomes "
        "(accepted or rejected). Review statuses: needs_review."
    )


def test_handle_api_base_url_change_clears_transient_backend_state() -> None:
    fake_streamlit, functions = load_streamlit_functions("reset_flow_state", "handle_api_base_url_change")
    handle_api_base_url_change = functions[-1]

    fake_streamlit.session_state.update(
        {
            "api_base_url": "http://127.0.0.1:8001/",
            "active_api_base_url": "http://127.0.0.1:8000",
            "upload_response": {"source": {"dataset_id": "src-1"}},
            "mapping_response": {"mappings": [{"source": "cust_id"}]},
            "codegen_response": {"language": "sql-dbt"},
            "runtime_config_snapshot": {"app_version": "0.1.0"},
            "runtime_config_snapshot_refreshed_at": 123.0,
            "admin_requirement": {"reachable": True, "requires_token": False},
            "backend_reachable": True,
            "backend_reachable_base_url": "http://127.0.0.1:8000",
            "catalog_results": [{"mapping_set_id": 771}],
            "benchmark_datasets": [{"dataset_id": 1}],
            "debug_runtime_config": {"app_version": "0.1.0"},
            "pending_workspace_section": "Output",
            "pending_governance_section": "Canonical",
            "review_focus_sources": ["KUNNR"],
            "active_top_level_area": "Catalog",
            "active_workspace_section": "Output",
        }
    )

    handle_api_base_url_change()

    assert fake_streamlit.session_state["api_base_url"] == "http://127.0.0.1:8001"
    assert fake_streamlit.session_state["active_api_base_url"] == "http://127.0.0.1:8001"
    assert fake_streamlit.session_state["pending_top_level_area"] == "Workspace"
    assert fake_streamlit.session_state["pending_workspace_section"] == "Setup"
    assert fake_streamlit.session_state["active_top_level_area"] == "Workspace"
    assert fake_streamlit.session_state["active_workspace_section"] == "Setup"
    assert "upload_response" not in fake_streamlit.session_state
    assert "mapping_response" not in fake_streamlit.session_state
    assert "codegen_response" not in fake_streamlit.session_state
    assert "runtime_config_snapshot" not in fake_streamlit.session_state
    assert "admin_requirement" not in fake_streamlit.session_state
    assert "backend_reachable" not in fake_streamlit.session_state
    assert "catalog_results" not in fake_streamlit.session_state
    assert "benchmark_datasets" not in fake_streamlit.session_state
    assert "debug_runtime_config" not in fake_streamlit.session_state
    assert "review_focus_sources" not in fake_streamlit.session_state
    assert fake_streamlit.session_state["last_action"]["level"] == "info"
    assert "http://127.0.0.1:8001" in fake_streamlit.session_state["last_action"]["message"]


def test_materialize_transformation_template_replaces_source_and_target_placeholders() -> None:
    _, [materialize_transformation_template] = load_streamlit_functions("materialize_transformation_template")

    rendered = materialize_transformation_template(
        {
            "code_template": 'df_target["{target}"] = df_source["{source}"].astype(str).str.strip()'
        },
        "legacy_name",
        "customer_name",
    )

    assert rendered == 'df_target["customer_name"] = df_source["legacy_name"].astype(str).str.strip()'


def test_trust_layer_rows_preserves_selected_candidate_signals() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "suggested_mapping_by_source",
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "transformation_mode",
        "canonical_concept_labels",
        "trust_layer_rows",
    )
    trust_layer_rows = functions[-1]

    fake_streamlit.session_state.update({"mapping_editor_state": {}})

    mapping_response = {
        "mappings": [
            {
                "source": "KUNNR",
                "target": "customer_id",
                "confidence": 0.94,
                "explanation": ["Internal metadata dictionary aligns both fields to concept 'Customer ID'."],
                "signals": {"knowledge": 1.0, "pattern": 1.0, "semantic": 0.9},
            }
        ],
        "ranked_mappings": [
            {
                "source": "KUNNR",
                "candidates": [
                    {
                        "target": "customer_id",
                        "confidence": 0.94,
                        "explanation": ["Internal metadata dictionary aligns both fields to concept 'Customer ID'."],
                        "canonical_details": {
                            "shared_concepts": [
                                {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 1.0}
                            ]
                        },
                        "signals": {"knowledge": 1.0, "pattern": 1.0, "semantic": 0.9},
                    }
                ],
            }
        ],
    }

    rows = trust_layer_rows(mapping_response)

    assert rows[0]["signals"]["knowledge"] == 1.0
    assert rows[0]["signals"]["pattern"] == 1.0
    assert rows[0]["canonical_details"]["shared_concepts"][0]["concept_id"] == "customer.id"


def test_trust_layer_rows_prefer_final_selected_row_metadata_when_target_matches() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "suggested_mapping_by_source",
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "transformation_mode",
        "trust_layer_rows",
    )
    trust_layer_rows = functions[-1]

    fake_streamlit.session_state.update({"mapping_editor_state": {}})

    mapping_response = {
        "mappings": [
            {
                "source": "purchaser",
                "target": "customer_id",
                "confidence": 0.4334,
                "explanation": [
                    "LLM validator preferred 'customer', but global one-to-one assignment selected this target instead."
                ],
                "signals": {"pattern": 1.0, "statistical": 0.96},
                "llm_consulted": True,
                "llm_recommendation": {
                    "selected_target": "customer",
                    "confidence": 0.5723,
                    "reasoning": ["The source field 'purchaser' represents a company name."],
                },
            }
        ],
        "ranked_mappings": [
            {
                "source": "purchaser",
                "candidates": [
                    {
                        "target": "customer_id",
                        "confidence": 0.4334,
                        "explanation": ["Strong pattern alignment: source text matches target text."],
                        "signals": {"pattern": 1.0, "statistical": 0.96},
                        "canonical_details": {"shared_concepts": []},
                    }
                ],
            }
        ],
    }

    rows = trust_layer_rows(mapping_response)

    assert rows[0]["explanation"] == [
        "LLM validator preferred 'customer', but global one-to-one assignment selected this target instead."
    ]
    assert rows[0]["llm_consulted"] is True
    assert rows[0]["llm_recommendation"]["selected_target"] == "customer"


def test_canonical_concept_labels_prefers_shared_concepts() -> None:
    _, [canonical_concept_labels] = load_streamlit_functions("canonical_concept_labels")

    labels = canonical_concept_labels(
        {
            "source_concepts": [{"concept_id": "customer.id", "display_name": "Customer ID"}],
            "target_concepts": [{"concept_id": "customer.id", "display_name": "Customer ID"}],
            "shared_concepts": [{"concept_id": "customer.id", "display_name": "Customer ID"}],
        }
    )

    assert labels == ["Customer ID (customer.id)"]


def test_canonical_path_label_formats_source_concept_target() -> None:
    _, functions = load_streamlit_functions("canonical_concept_labels", "canonical_path_label")
    canonical_path_label = functions[-1]

    label = canonical_path_label(
        "sold_to_party",
        "customer_id",
        {
            "shared_concepts": [{"concept_id": "customer.id", "display_name": "Customer ID"}],
        },
    )

    assert label == "sold_to_party -> Customer ID (customer.id) -> customer_id"


def test_current_mapping_rows_includes_canonical_path_for_selected_candidate() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "suggested_mapping_by_source",
        "validator_badge",
        "canonical_concept_labels",
        "canonical_path_label",
        "current_mapping_rows",
    )
    current_mapping_rows = functions[-1]

    fake_streamlit.session_state.update({"mapping_editor_state": {}})

    rows = current_mapping_rows(
        {
            "mappings": [
                {
                    "source": "sold_to_party",
                    "target": "customer_id",
                    "confidence": 0.94,
                    "confidence_label": "high_confidence",
                    "status": "accepted",
                    "method": "multi_signal_heuristic",
                    "canonical_details": {
                        "shared_concepts": [
                            {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 1.0}
                        ]
                    },
                }
            ],
            "ranked_mappings": [
                {
                    "source": "sold_to_party",
                    "candidates": [
                        {
                            "target": "customer_id",
                            "confidence": 0.94,
                            "confidence_label": "high_confidence",
                            "method": "multi_signal_heuristic",
                            "canonical_details": {
                                "shared_concepts": [
                                    {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 1.0}
                                ]
                            },
                        }
                    ],
                }
            ],
        }
    )

    assert rows[0]["canonical_path"] == "sold_to_party -> Customer ID (customer.id) -> customer_id"
    assert rows[0]["canonical_status"] == "shared_match"
    assert rows[0]["shared_concepts"] == "Customer ID (customer.id)"
    assert rows[0]["source_concepts"] == ""
    assert rows[0]["target_concepts"] == ""


def test_current_mapping_rows_marks_source_only_canonical_mismatch() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "suggested_mapping_by_source",
        "validator_badge",
        "canonical_concept_labels",
        "canonical_path_label",
        "current_mapping_rows",
    )
    current_mapping_rows = functions[-1]

    fake_streamlit.session_state.update({"mapping_editor_state": {}})

    rows = current_mapping_rows(
        {
            "mappings": [
                {
                    "source": "REGIO",
                    "target": "region_code",
                    "confidence": 0.57,
                    "confidence_label": "low_confidence",
                    "status": "needs_review",
                    "method": "multi_signal_heuristic",
                    "canonical_details": {
                        "source_concepts": [
                            {"concept_id": "supplier.region_code", "display_name": "Supplier Region Code", "strength": 0.8}
                        ],
                        "target_concepts": [],
                        "shared_concepts": [],
                    },
                }
            ],
            "ranked_mappings": [
                {
                    "source": "REGIO",
                    "candidates": [
                        {
                            "target": "region_code",
                            "confidence": 0.57,
                            "confidence_label": "low_confidence",
                            "method": "multi_signal_heuristic",
                            "canonical_details": {
                                "source_concepts": [
                                    {"concept_id": "supplier.region_code", "display_name": "Supplier Region Code", "strength": 0.8}
                                ],
                                "target_concepts": [],
                                "shared_concepts": [],
                            },
                        }
                    ],
                }
            ],
        }
    )

    assert rows[0]["canonical_status"] == "source_only"
    assert rows[0]["canonical_path"] == "REGIO -> Supplier Region Code (supplier.region_code) -> region_code"
    assert rows[0]["source_concepts"] == "Supplier Region Code (supplier.region_code)"
    assert rows[0]["target_concepts"] == ""
    assert rows[0]["shared_concepts"] == ""


def test_default_editor_entry_preserves_unmapped_selected_result() -> None:
    _, [default_editor_entry] = load_streamlit_functions("default_editor_entry")

    entry = default_editor_entry(
        {
            "selected": {
                "target": None,
                "status": "needs_review",
            },
            "candidates": [
                {
                    "target": "customer.phone",
                }
            ],
        }
    )

    assert entry["target"] == ""
    assert entry["status"] == "needs_review"
    assert entry["suggested_target"] == ""
    assert entry["manual_canonical_concept"] == ""


def test_default_editor_entry_initializes_manual_canonical_concept_blank() -> None:
    _, [default_editor_entry] = load_streamlit_functions("default_editor_entry")

    entry = default_editor_entry(
        {
            "selected": {
                "target": "customer_id",
                "status": "accepted",
                "resolution_type": "direct_mapping",
                "resolution_payload": {},
            },
            "candidates": [
                {
                    "target": "customer_id",
                }
            ],
        }
    )

    assert entry["manual_canonical_concept"] == ""


def test_selected_target_options_include_unmapped_for_no_match_selection() -> None:
    _, [selected_target_options] = load_streamlit_functions("selected_target_options")

    options = selected_target_options(
        {
            "selected": {
                "target": None,
                "status": "needs_review",
            },
            "candidates": [
                {"target": "customer.phone"},
                {"target": "address.postal_code"},
            ],
        }
    )

    assert options == ["", "customer.phone", "address.postal_code"]


def test_current_mapping_rows_preserves_unmapped_no_match_selection() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "suggested_mapping_by_source",
        "validator_badge",
        "canonical_concept_labels",
        "canonical_path_label",
        "current_mapping_rows",
    )
    current_mapping_rows = functions[-1]

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "LAND1": {
                    "target": "",
                    "status": "needs_review",
                }
            }
        }
    )

    rows = current_mapping_rows(
        {
            "mappings": [
                {
                    "source": "LAND1",
                    "target": None,
                    "confidence": 0.0,
                    "confidence_label": "low_confidence",
                    "status": "needs_review",
                    "method": "llm_validator_no_match",
                    "canonical_details": {
                        "shared_concepts": [],
                    },
                }
            ],
            "ranked_mappings": [
                {
                    "source": "LAND1",
                    "candidates": [
                        {
                            "target": "customer.phone",
                            "confidence": 0.29,
                            "confidence_label": "low_confidence",
                            "method": "multi_signal_heuristic",
                            "canonical_details": {
                                "shared_concepts": [],
                            },
                        }
                    ],
                }
            ],
        }
    )

    assert rows[0]["target"] == ""
    assert rows[0]["confidence"] == 0.0
    assert rows[0]["validator"] == "Llm Validator No Match"


def test_canonical_concept_groups_group_selected_mappings_by_shared_concept() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "suggested_mapping_by_source",
        "canonical_concept_labels",
        "canonical_path_label",
        "canonical_concept_groups",
    )
    canonical_concept_groups = functions[-1]

    fake_streamlit.session_state.update({"mapping_editor_state": {}})

    groups = canonical_concept_groups(
        {
            "mappings": [
                {
                    "source": "sold_to_party",
                    "target": "customer_id",
                    "canonical_details": {
                        "shared_concepts": [
                            {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 1.0}
                        ]
                    },
                },
                {
                    "source": "ship_to_party",
                    "target": "customer_id",
                    "canonical_details": {
                        "shared_concepts": [
                            {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 0.9}
                        ]
                    },
                },
                {
                    "source": "client_mail",
                    "target": "customer_email",
                    "canonical_details": {
                        "shared_concepts": [
                            {"concept_id": "customer.email", "display_name": "Customer Email", "strength": 1.0}
                        ]
                    },
                },
            ],
            "ranked_mappings": [
                {
                    "source": "sold_to_party",
                    "candidates": [
                        {
                            "target": "customer_id",
                            "canonical_details": {
                                "shared_concepts": [
                                    {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 1.0}
                                ]
                            },
                        }
                    ],
                },
                {
                    "source": "ship_to_party",
                    "candidates": [
                        {
                            "target": "customer_id",
                            "canonical_details": {
                                "shared_concepts": [
                                    {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 0.9}
                                ]
                            },
                        }
                    ],
                },
                {
                    "source": "client_mail",
                    "candidates": [
                        {
                            "target": "customer_email",
                            "canonical_details": {
                                "shared_concepts": [
                                    {"concept_id": "customer.email", "display_name": "Customer Email", "strength": 1.0}
                                ]
                            },
                        }
                    ],
                },
            ],
        }
    )

    assert groups == [
        {
            "concept": "Customer Email",
            "concept_id": "customer.email",
            "mapping_count": 1,
            "source_columns": "client_mail",
            "target_columns": "customer_email",
            "canonical_paths": "client_mail -> Customer Email (customer.email) -> customer_email",
        },
        {
            "concept": "Customer ID",
            "concept_id": "customer.id",
            "mapping_count": 2,
            "source_columns": "ship_to_party | sold_to_party",
            "target_columns": "customer_id",
            "canonical_paths": "ship_to_party -> Customer ID (customer.id) -> customer_id | sold_to_party -> Customer ID (customer.id) -> customer_id",
        },
    ]


def test_source_concept_rows_lists_selected_source_to_concept_paths() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "suggested_mapping_by_source",
        "source_concept_rows",
    )
    source_concept_rows = functions[-1]

    fake_streamlit.session_state.update({"mapping_editor_state": {}})

    rows = source_concept_rows(
        {
            "mappings": [
                {
                    "source": "sold_to_party",
                    "target": "customer_id",
                    "canonical_details": {
                        "shared_concepts": [
                            {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 1.0}
                        ]
                    },
                }
            ],
            "ranked_mappings": [
                {
                    "source": "sold_to_party",
                    "candidates": [
                        {
                            "target": "customer_id",
                            "canonical_details": {
                                "shared_concepts": [
                                    {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 1.0}
                                ]
                            },
                        }
                    ],
                }
            ],
        }
    )

    assert rows == [
        {
            "source": "sold_to_party",
            "concept": "Customer ID",
            "concept_id": "customer.id",
            "target": "customer_id",
            "strength": 1.0,
        }
    ]


def test_concept_target_rows_groups_selected_targets_by_concept() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "suggested_mapping_by_source",
        "concept_target_rows",
    )
    concept_target_rows = functions[-1]

    fake_streamlit.session_state.update({"mapping_editor_state": {}})

    rows = concept_target_rows(
        {
            "mappings": [
                {
                    "source": "sold_to_party",
                    "target": "customer_id",
                    "canonical_details": {
                        "shared_concepts": [
                            {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 1.0}
                        ]
                    },
                },
                {
                    "source": "ship_to_party",
                    "target": "customer_id",
                    "canonical_details": {
                        "shared_concepts": [
                            {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 0.9}
                        ]
                    },
                },
            ],
            "ranked_mappings": [
                {
                    "source": "sold_to_party",
                    "candidates": [
                        {
                            "target": "customer_id",
                            "canonical_details": {
                                "shared_concepts": [
                                    {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 1.0}
                                ]
                            },
                        }
                    ],
                },
                {
                    "source": "ship_to_party",
                    "candidates": [
                        {
                            "target": "customer_id",
                            "canonical_details": {
                                "shared_concepts": [
                                    {"concept_id": "customer.id", "display_name": "Customer ID", "strength": 0.9}
                                ]
                            },
                        }
                    ],
                },
            ],
        }
    )

    assert rows == [
        {
            "concept": "Customer ID",
            "concept_id": "customer.id",
            "target": "customer_id",
            "source_columns": "ship_to_party | sold_to_party",
            "mapping_count": 2,
        }
    ]


def test_has_knowledge_match_uses_signal_when_present() -> None:
    _, [has_knowledge_match] = load_streamlit_functions("has_knowledge_match")

    assert has_knowledge_match({"knowledge": 1.0}, None) is True
    assert has_knowledge_match({"knowledge": 0.0}, None) is False


def test_has_knowledge_match_falls_back_to_explanation_text() -> None:
    _, [has_knowledge_match] = load_streamlit_functions("has_knowledge_match")

    assert has_knowledge_match(
        {},
        ["Context prior: source SAP KNA1.KUNNR aligns with target Workday Customer.Customer_ID."],
    ) is True
    assert has_knowledge_match({}, ["Pattern similarity is strong."]) is False


def test_has_canonical_match_uses_signal_when_present() -> None:
    _, [has_canonical_match] = load_streamlit_functions("has_canonical_match")

    assert has_canonical_match({"canonical": 1.0}, None) is True
    assert has_canonical_match({"canonical": 0.0}, None) is False


def test_has_canonical_match_falls_back_to_explanation_text() -> None:
    _, [has_canonical_match] = load_streamlit_functions("has_canonical_match")

    assert has_canonical_match(
        {},
        ["Canonical glossary aligns both fields to business concept 'Customer ID' (customer.id)."],
    ) is True
    assert has_canonical_match({}, ["Pattern similarity is strong."]) is False


def test_canonical_explanation_lines_extracts_only_canonical_messages() -> None:
    _, [canonical_explanation_lines] = load_streamlit_functions("canonical_explanation_lines")

    lines = canonical_explanation_lines(
        [
            "Canonical glossary aligns both fields to business concept 'Customer ID' (customer.id).",
            "Field names are lexically very similar.",
        ]
    )

    assert lines == ["Canonical glossary aligns both fields to business concept 'Customer ID' (customer.id)."]


def test_build_mapping_set_payload_uses_current_dataset_ids_and_decisions() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "build_mapping_decisions",
        "build_mapping_set_payload",
    )
    build_mapping_set_payload = functions[-1]

    fake_streamlit.session_state.update(
        {
            "upload_response": {
                "mapping_mode": "canonical",
                "source": {"dataset_id": "source-1"},
                "target_system": "canonical",
            },
            "mapping_response": {
                "canonical_coverage": {
                    "source": {"unmatched_columns": ["LAND1"]},
                    "project": {"concepts": ["customer.id"]},
                }
            },
            "mapping_editor_state": {
                "cust_id": {"target": "customer_id", "status": "accepted"},
            },
            "manual_transform_cust_id": "",
            "manual_apply_cust_id": False,
        }
    )

    payload = build_mapping_set_payload(
        "customer-master",
        "ba-user",
        "Initial draft",
        "governance-team",
        "analyst-1",
        "Ready for review",
    )

    assert payload == {
        "name": "customer-master",
        "source_dataset_id": "source-1",
        "target_dataset_id": None,
        "mapping_decisions": [
            {
                "source": "cust_id",
                "target": "customer_id",
                "status": "accepted",
                "resolution_type": "direct_mapping",
            }
        ],
        "integration_name": "customer-master",
        "target_system": "canonical",
        "artifact_type": "canonical-only",
        "canonical_concepts": ["customer.id"],
        "unmatched_sources": ["LAND1"],
        "created_by": "ba-user",
        "note": "Initial draft",
        "owner": "governance-team",
        "assignee": "analyst-1",
        "review_note": "Ready for review",
    }


def test_export_mapping_excel_bytes_creates_tabular_mapping_workbook() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "build_mapping_decisions",
        "export_mapping_excel_bytes",
    )
    export_mapping_excel_bytes = functions[-1]

    fake_streamlit.session_state.update(
        {
            "mapping_editor_state": {
                "cust_id": {
                    "target": "customer_id",
                    "status": "accepted",
                    "manual_transformation_code": 'df_source["cust_id"].astype(str)',
                    "manual_apply_transformation": True,
                },
                "phone": {
                    "target": "phone_number",
                    "status": "needs_review",
                },
            },
            "manual_transform_cust_id": 'df_source["cust_id"].astype(str)',
            "manual_apply_cust_id": True,
            "manual_transform_phone": "",
            "manual_apply_phone": False,
        }
    )

    payload = export_mapping_excel_bytes()
    workbook = load_workbook(BytesIO(payload))
    worksheet = workbook["mapping_decisions"]

    assert worksheet.max_row == 3
    assert worksheet.max_column == 6
    assert [cell.value for cell in worksheet[1]] == ["source", "target", "status", "resolution_type", "resolution_payload", "transformation_code"]
    assert [cell.value for cell in worksheet[2]] == [
        "cust_id",
        "customer_id",
        "accepted",
        "direct_mapping",
        None,
        'df_source["cust_id"].astype(str)',
    ]
    assert [cell.value for cell in worksheet[3]] == ["phone", "phone_number", "needs_review", "direct_mapping", None, None]


def test_export_mapping_payload_includes_mapping_decision_audit() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "resolve_suggested_transformation_code",
        "effective_transformation_code",
        "build_mapping_decisions",
        "export_mapping_payload",
    )
    export_mapping_payload = functions[-1]

    fake_streamlit.session_state.update(
        {
            "upload_response": {
                "source": {"dataset_id": "source-1"},
                "target": {"dataset_id": "target-1"},
            },
            "mapping_editor_state": {
                "cust_id": {"target": "customer_id", "status": "accepted"},
            },
            "manual_transform_cust_id": "",
            "manual_apply_cust_id": False,
            "mapping_decision_audit": {
                "cust_id": {
                    "origin": "llm_proposal",
                    "applied_at": "2026-05-22T12:00:00+00:00",
                    "details": {"mode": "accept_current"},
                }
            },
        }
    )

    payload = json.loads(export_mapping_payload())

    assert payload["mapping_decision_audit"] == {
        "cust_id": {
            "origin": "llm_proposal",
            "applied_at": "2026-05-22T12:00:00+00:00",
            "details": {"mode": "accept_current"},
        }
    }


def test_apply_imported_mapping_payload_restores_mapping_decision_audit() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "schema_column_names",
        "apply_imported_mapping_payload",
    )
    apply_imported_mapping_payload = functions[-1]

    fake_streamlit.session_state.update(
        {
            "upload_response": {
                "source": {
                    "schema_profile": {
                        "columns": [
                            {"name": "cust_id"},
                            {"name": "phone"},
                        ]
                    }
                }
            },
            "mapping_editor_state": {},
            "mapping_decision_audit": {
                "stale_field": {
                    "origin": "manual_mapping",
                    "applied_at": "old",
                    "details": {},
                }
            },
        }
    )

    payload = {
        "mapping_decisions": [
            {"source": "cust_id", "target": "customer_id", "status": "accepted"},
        ],
        "mapping_decision_audit": {
            "cust_id": {
                "origin": "llm_proposal",
                "applied_at": "2026-05-22T12:00:00+00:00",
                "details": {"mode": "switch_target", "confidence": 0.91},
            },
            "unknown_source": {
                "origin": "llm_proposal",
                "applied_at": "2026-05-22T12:00:00+00:00",
                "details": {"mode": "switch_target"},
            },
        },
    }

    apply_imported_mapping_payload(json.dumps(payload).encode("utf-8"))

    assert fake_streamlit.session_state["mapping_decision_audit"] == {
        "cust_id": {
            "origin": "llm_proposal",
            "applied_at": "2026-05-22T12:00:00+00:00",
            "details": {"mode": "switch_target", "confidence": 0.91},
        }
    }


def test_apply_imported_mapping_payload_normalizes_null_transformation_code() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "schema_column_names",
        "apply_imported_mapping_payload",
    )
    apply_imported_mapping_payload = functions[-1]

    fake_streamlit.session_state.update(
        {
            "upload_response": {
                "source": {
                    "schema_profile": {
                        "columns": [
                            {"name": "cust_id"},
                        ]
                    }
                }
            },
            "mapping_editor_state": {},
        }
    )

    payload = {
        "mapping_decisions": [
            {"source": "cust_id", "target": "customer_id", "status": "accepted", "transformation_code": None},
        ]
    }

    apply_imported_mapping_payload(json.dumps(payload).encode("utf-8"))

    assert fake_streamlit.session_state["mapping_editor_state"]["cust_id"]["manual_transformation_code"] == ""
    assert fake_streamlit.session_state["mapping_editor_state"]["cust_id"]["resolution_type"] == "direct_mapping"
    assert fake_streamlit.session_state["mapping_editor_state"]["cust_id"]["resolution_payload"] == {}
    assert fake_streamlit.session_state["manual_transform_cust_id"] == ""
    assert fake_streamlit.session_state["manual_apply_cust_id"] is False


def test_apply_imported_mapping_payload_preserves_resolution_type() -> None:
    fake_streamlit, functions = load_streamlit_functions(
        "schema_column_names",
        "apply_imported_mapping_payload",
    )
    apply_imported_mapping_payload = functions[-1]

    fake_streamlit.session_state.update(
        {
            "upload_response": {
                "source": {
                    "schema_profile": {
                        "columns": [
                            {"name": "op_type"},
                        ]
                    }
                }
            },
            "mapping_editor_state": {},
        }
    )

    payload = {
        "mapping_decisions": [
            {
                "source": "op_type",
                "target": "operation_category",
                "status": "accepted",
                "resolution_type": "fixed_value",
                "resolution_payload": {"value": "CUSTOMER"},
            },
        ]
    }

    apply_imported_mapping_payload(json.dumps(payload).encode("utf-8"))

    assert fake_streamlit.session_state["mapping_editor_state"]["op_type"]["resolution_type"] == "fixed_value"
    assert fake_streamlit.session_state["mapping_editor_state"]["op_type"]["resolution_payload"] == {"value": "CUSTOMER"}